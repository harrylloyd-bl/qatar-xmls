# -*- coding: utf-8 -*-
"""
Created on Thu Jan 11 10:35:56 2024

@author: amilighe

The code should be stored in the same folder as the gather sheet. The output
will also be in the same folder.

Notes on the creation of the template:
    - Authority files should be in the format: Name>role>altrender|Name>role>altreder
      with no extra spaces. The authority file sheet's name is hard-coded in the code.
    - The authority file sheet should be updated to match the authority files
      in each item to Gather. If a new sheet is created, update the name of the
      sheet in line 144.
    - The IAMS template sheet should have one tab for each item and one row for each child.
    - When inputting the IAMS sheet's name, omit the .xlsx as it will be added in the code
    - Note that Authority file processing takes a longer time as it needs to
      reference a high volume of items.
    - The code only has very basic error control. Always check the output.
"""

from datetime import datetime

from lxml import etree
from lxml.builder import ElementMaker
from lxml.etree import Comment
from openpyxl import load_workbook


# Definitions used to create the nodes:

def get_header(ws):
    header = []
    for cell in ws[1]:  # don't reference a variable in the main script inside a function, make it an argument of the function
        header.append(cell.value)
    return header

# use two (or at least one) blank lines between functions
def start_record(rec_num):  # consistent use of underscore ('snake_case') in function definitions
    return {"StartRecord": rec_num}


def tid(row, arg):
    global tid_num  # TODO add docs to this fn to explain why it keeps track of tid_num
    if row[arg].value != None:  # use is not None instead of != None (Python style guide)
        fixed_shelfmark = row[5].value.replace("/", "_").replace(" ", "_").replace("-", "_").replace(",", "_")
        tid_full = fixed_shelfmark + "_" + str(tid_num)
        tid_num = tid_num + 1
        return {"tid": tid_full}
    else:
        return ""


def content(row, arg):
    if row[arg].value:
        return row[arg].value
    else:
        return ""


def labels(row, arg, label):
    content = str(row[arg].value)
    return {label: content}


def header_label(header_row, arg, label):
    label_title = header_row[arg]
    return {label: label_title}


def date_in_full(row):
    if row[16].value:
        date_item = str(row[15].value) + "-" + str(row[16].value)
    else:
        date_item = str(row[15].value)
    return date_item


def pcontent(row, arg):
    content = []
    if row[arg].value:
        lines = row[arg].value.split("\n")
        for line in lines:
            p = E.p(line, tid(row, arg))
            content.append(p)
    else:
        p = E.p("")
        content.append(p)
    return content


# Authority Files processing definitions
def search_auth_file(line, auth_ws):  # This is how auth_ws should be referenced, as an arg in the fn. Compare to how it's referenced in authority_files()
    for row_num, row in enumerate(auth_ws.iter_rows()):
        cell = row[0]
        if str(cell.value).strip().lower() == str(line).strip().lower():
            return row_num + 1
    return "not_found"


def auth_dets(arg, label):
    if arg == "not_allocated":
        return {}
    else:
        return {label: arg}


def authfilenumber(auth_row_num):  # consider changing to auth_filenumber/auth_file_number for consistency with convenction
    if auth_row_num != "not_found":
        auth_row = auth_ws[auth_row_num]  # TODO auth_ws should be an arg
        return {"authfilenumber": auth_row[18].value}
    else:
        return {"authfilenumber": "not_found"}


def authority_files(row, arg):  # add underscore to match convention
    if row[arg].value:
        lines = row[arg].value.split("|")
        full_text = []
        for line in lines:
            attributes = line.split(">")
            subject = attributes[0]
            role_type = "not_allocated"  # these are defaults, so can define first
            altrender_type = "not_allocated"
            if len(attributes) > 1 and attributes[1]:
                role_type = attributes[1]
            if len(attributes) > 2:
                altrender_type = attributes[2]

            auth_row_num = search_auth_file(subject, auth_ws)  # TODO auth_ws should be an arg in authority_files()
            # Reorder these 48/49/50/51/52 unless this order makes more sense with tag types
            # or save writing out the arguments each time with:
            # element_dict = {48: E.corpname, 49: E.persname, 50: E.famname, 51: E.geogname, 52: E.subject}
            # text = element_dict[arg](subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),
            #                         {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            if arg == 50:  # TODO I think E should be an argument in authority_files as well
                text = E.corpname(subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),  # line break for readability
                                  {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            elif arg == 48:
                text = E.persname(subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),
                                  {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            elif arg == 49:
                text = E.famname(subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),
                                 {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            elif arg == 51:
                text = E.geogname(subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),
                                  {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            elif arg == 52:
                text = E.subject(subject, authfilenumber(auth_row_num), auth_dets(role_type, "role"),
                                 {"source": "IAMS"}, auth_dets(altrender_type, "altrender"), tid(row, arg))
            full_text.append(text)
        return full_text
    else:
        return ""


# The actual code starts here: the input should be the name of the spreadsheet to gather.
# One tab each shelfmark to gather.

wb_input = input('Please write the name of the spreadsheet to Gather. Omit ".xlsx": ')
wb_name = wb_input + '.xlsx'
wb = load_workbook(wb_name, read_only=True)
shelfmarks = wb.sheetnames
for shelfmark_modified in shelfmarks:  # Use sm instead of shelfmark_modified for concision?
    # Do you want the script to continue if you don't find the sheet or do you want to raise the KeyError
    # ws will be undefined in `for row in ws.iter_rows` if you don't raise the KeyError
    try:
        ws = wb[shelfmark_modified]
    except KeyError:
        print("Sheet not found")

    # This part defines where the authority files details are held.

    auth_file_name = 'Authorities_combined.xlsx'  # TODO auth_ws only needs to be defined once, so can do before for loop
    auth_file_wb = load_workbook(auth_file_name, read_only=True)
    try:
        auth_ws = auth_file_wb["1"]
    except KeyError:
        print("Sheet not found")

    # This is where the nodes are established. If a node has the wrong name, change it in E.node
    # TODO add a <?xml version="1.0" encoding="utf-8" ?> tag?
    E = ElementMaker(namespace="urn:isbn:1-931666-22-9",
                     nsmap={'ead': "urn:isbn:1-931666-22-9", 'xlink': "http://www.w3.org/1999/xlink",
                            'xsi': "http://www.w3.org/2001/XMLSchema-instance"})
    # TODO does this actually help readability - or would it be better to just use E.ead()
    EAD = E.ead
    EADHEADER = E.eadheader
    EADID = E.eadid
    FILEDESC = E.filedesc
    TITLESTMT = E.titlestmt
    TITLEPROPER = E.titleproper
    PROFILEDESC = E.profiledesc
    CREATION = E.creation
    DATE = E.date
    LANGUSAGE = E.langusage
    LANGUAGE = E.language
    ARCHDESC = E.archdesc
    DID = E.did
    REPOSITORY = E.repository
    UNITID = E.unitid
    UNITTITLE = E.unittitle
    TITLE = E.title
    UNITDATE = E.unitdate
    LANGMATERIAL = E.langmaterial
    PHYSDESC = E.physdesc
    EXTENT = E.extent
    ACCESSRESTRICT = E.accessrestrict
    P = E.p
    LEGALSTATUS = E.legalstatus
    ACCRUALS = E.accruals
    BIOGHIST = E.bioghist
    APPRAISAL = E.appraisal
    ARRANGEMENT = E.arrangement
    PHYSTECH = E.phystech
    SCOPECONTENT = E.scopecontent
    LIST = E.list
    USERRESTRICT = E.userrestrict  # TODO should this be userestrict with only one r? This is how it appears in IOR_L_PS_5_270__ff_197_200_original.xml
    ODD = E.odd
    CONTROLACCESS = E.controlaccess
    GENREFORM = E.genreform
    PERSNAME = E.persname
    FAMNAME = E.famname
    CORPNAME = E.corpname
    SUBJECT = E.subject
    GEOGNAME = E.geogname
    NOTE = E.note

    full_ead = EAD()

    rec_num = 1  # Move rec_num/tid_num defs here to be closer to where they're used
    tid_num = 1  # TODO explain that this is updated every time tid() is called?
    # This part creates the tree for each child shelfmark.
    for row in ws.iter_rows(min_row=2, values_only=False):
        ead = EAD()
        comment = Comment(f"New record starts here {row[5].value}")
        full_ead.append(comment)
        shelfmark = str(row[5].value)
        print(shelfmark)
        header_row = get_header(ws)  # TODO doesn't change between loops, can define before `for row in ws.iter_rows()..`

        # header
        # consider eadheader = E.eadheader(start_record(str(rec_num))))
        eadheader = EADHEADER(start_record(str(rec_num)))
        ead.append(eadheader)

        eadid = EADID(str(shelfmark), tid(row, 5))
        eadheader.append(eadid)

        filedesc = FILEDESC()  # wrapper node, should not have info
        eadheader.append(filedesc)

        titlestmt = TITLESTMT()  # wrapper node, should not have info
        filedesc.append(titlestmt)

        titleproper = TITLEPROPER()  # as far as I can see, not used in IAMS material
        titlestmt.append(titleproper)

        profiledesc = PROFILEDESC()  # wrapper node, should not have info
        eadheader.append(profiledesc)

        creation = CREATION()  # this has to do with finding aids, not used in Qatar(?)
        profiledesc.append(creation)

        # TODO I'd define these two `date` variables with different names, even though their tag in the xml is the same
        date = DATE(str(datetime.now().strftime("%Y-%m-%dT%H:%M:%S")), {"type": "exported"}, tid(row, 5))
        creation.append(date)

        date = DATE(str(wb.properties.modified.strftime("%Y-%m-%dT%H:%M:%S")), {"type": "modified"}, tid(row, 5))
        creation.append(date)

        langusage = LANGUSAGE()  # as far as I can see, not used in IAMS material
        profiledesc.append(langusage)

        language = LANGUAGE(content(row, 44), labels(row, 45, "langcode"), labels(row, 47, "scriptcode"), tid(row, 40))  # this is language of the description
        langusage.append(language)

        # archdesc
        archdesc = ARCHDESC(labels(row, 4, "level"))
        ead.append(archdesc)

        did = DID()  # wrapper node, should not have info
        archdesc.append(did)

        repository = REPOSITORY(row[0].value + ": " + row[1].value, tid(row, 0))  # British Library: Indian Office Records
        did.append(repository)

        unitid = UNITID(shelfmark, {"label": "IAMS_label_NA"}, {"identifier": "ark_identifier"}, tid(row, 5))  # These are the IAMS identifiers (ark and number)
        did.append(unitid)

        unittitle = UNITTITLE(header_label(header_row, 10, "label"))  # this will say "title"
        did.append(unittitle)

        title = TITLE(content(row, 10), tid(row, 10))  # Item title
        unittitle.append(title)

        unittitle = UNITTITLE(content(row, 7), header_label(header_row, 7, "label"), tid(row, 7))
        did.append(unittitle)  # Former external reference

        unittitle = UNITTITLE(content(row, 6), header_label(header_row, 6, "label"), tid(row, 6))
        did.append(unittitle)  # Former internal reference

        unitdate = UNITDATE(date_in_full(row), {"datechar": "Creation"}, labels(row, 18, "calendar"),
                            labels(row, 17, "era"), labels(row, 14, "normal"), tid(row, 14))
        did.append(unitdate)  # Date of the material

        for r in ((40, 41), (42, 43)):  # Simplify code with for loop
            # TODO will need to explain separately that 40/41 are language and 42/43 are script
            langmaterial = LANGMATERIAL()  # This is language
            did.append(langmaterial)

            # This allows for multiple material languages and language codes separated by |
            languages = row[r[0]].value.split("|")
            lang_codes = row[r[1]].value.split("|")
            for l, c in zip(languages, lang_codes):
                language = LANGUAGE(l, {"langcode": c}, tid(row, 41))
                langmaterial.append(language)

        physdesc = PHYSDESC()  # wrapper node, should not have info
        did.append(physdesc)

        extent = EXTENT(content(row, 19), tid(row, 19))
        physdesc.append(extent)

        accessrestrict = ACCESSRESTRICT()
        for p in pcontent(row, 25):
            accessrestrict.append(p)
        archdesc.append(accessrestrict)

        accessrestrict = ACCESSRESTRICT()  # This second accessrestrict is a wrapper node
        archdesc.append(accessrestrict)

        legalstatus = LEGALSTATUS(content(row, 71), tid(row, 71))
        accessrestrict.append(legalstatus)

        accruals = ACCRUALS()
        for p in pcontent(row, 23):
            accruals.append(p)
        archdesc.append(accruals)

        bioghist = BIOGHIST()
        for p in pcontent(row, 24):
            bioghist.append(p)
        archdesc.append(bioghist)

        appraisal = APPRAISAL()
        for p in pcontent(row, 22):
            appraisal.append(p)
        archdesc.append(appraisal)

        arrangement = ARRANGEMENT()
        for p in pcontent(row, 31):
            arrangement.append(p)
        archdesc.append(arrangement)

        phystech = PHYSTECH()
        for p in pcontent(row, 21):
            phystech.append(p)
        archdesc.append(phystech)

        # This section allows for bullet points in the scope and content
        scopecontent = SCOPECONTENT()
        lists = LIST()
        if row[20].value.find("-"):
            list_content = []
            top_content = []
            bottom_content = []
            lines = row[20].value.split("\n")
            for line in lines:
                if line.startswith("-"):
                    list_content.append(line)
                else:  # TODO could you replace the else with elif not list_content: top_content.. then a final else: bottom_content... ?
                    if not list_content:
                        top_content.append(line)
                    else:
                        bottom_content.append(line)
            for section in top_content:
                p = E.p(section, tid(row, 20))
                scopecontent.append(p)
            for section in list_content:
                item = E.item(section.strip("-"), tid(row, 20))
                lists.append(item)
                scopecontent.append(lists)
            for section in bottom_content:
                p = E.p(section, tid(row, 20))
                scopecontent.append(p)
        else:
            for p in pcontent(row, 20):
                scopecontent.append(p)
        archdesc.append(scopecontent)

        userestrict = USERRESTRICT()
        for p in pcontent(row, 27):
            userestrict.append(p)
        archdesc.append(userestrict)

        odd = ODD()
        for p in pcontent(row, 36):  # unsure if this is the correct mapping for this field!
            odd.append(p)
        archdesc.append(odd)

        controlaccess = CONTROLACCESS()
        genreform = GENREFORM(content(row, 79), {"source": "IAMS"}, tid(row, 79))
        controlaccess.append(genreform)

        # Authority files processing starts here:
        print(f"{datetime.now().strftime('%Y-%m-%dT%H:%M:%S')} processing authority files for {shelfmark}...")  # use f string for clarity
        for arg in range(48, 54, 1):
            for authorityfile in authority_files(row, arg):  # shorten authorityfile to af ?
                controlaccess.append(authorityfile)
        archdesc.append(controlaccess)
        # End of authority files.

        note = NOTE(header_label(header_row, 2, "type"))
        for p in pcontent(row, 1):
            note.append(p)
        controlaccess.append(note)

        note = NOTE(header_label(header_row, 2, "type"))
        for p in pcontent(row, 2):
            note.append(p)
        controlaccess.append(note)

        rec_num = rec_num + 1

        # This part puts together the two parts of each tree (header+description)
        # This will append as many children as there are in the Excel tab
        full_ead.append(eadheader)
        full_ead.append(archdesc)

    # This part writes out the XML file
    with open(shelfmark_modified + '.xml', 'wb') as f:
        f.write(etree.tostring(full_ead, encoding='UTF-8', pretty_print=True))

    print(shelfmark + ' complete \n')

wb.close()  # TODO check - I don't think these .close() statements are necessary?
auth_file_wb.close()
print('Gather complete!')
